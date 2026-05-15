from fastapi import FastAPI, HTTPException
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from pydantic import BaseModel
from typing import Dict, Any
import pandas as pd
import re
import os
import shutil
from datetime import datetime
from openpyxl import load_workbook

import json

app = FastAPI()

DATA_DIR = None
ASSET_DIR = None
_CONFIG_FILE = os.path.join(os.path.dirname(__file__), ".last_folder")
_ASSET_CONFIG_FILE = os.path.join(os.path.dirname(__file__), ".last_asset_folder")

def _load_last_folder():
    global DATA_DIR
    if os.path.exists(_CONFIG_FILE):
        try:
            with open(_CONFIG_FILE, "r", encoding="utf-8") as f:
                path = f.read().strip()
            if path and os.path.isdir(path):
                DATA_DIR = path
        except Exception:
            pass

def _load_last_asset_folder():
    global ASSET_DIR
    if os.path.exists(_ASSET_CONFIG_FILE):
        try:
            with open(_ASSET_CONFIG_FILE, "r", encoding="utf-8") as f:
                path = f.read().strip()
            if path and os.path.isdir(path):
                ASSET_DIR = path
        except Exception:
            pass

_load_last_folder()
_load_last_asset_folder()


class FolderRequest(BaseModel):
    path: str


def find_header_and_data_start(df, key_field: str):
    header_row = None
    for i in range(min(10, len(df))):
        row_values = [str(v).strip().lower() for v in df.iloc[i].tolist() if pd.notna(v)]
        if key_field.lower() in row_values:
            header_row = i
            break

    if header_row is None:
        return None, None, None

    headers = df.iloc[header_row].tolist()
    col_map = {}
    for idx, h in enumerate(headers):
        h_str = str(h).strip() if pd.notna(h) else ""
        if h_str:
            col_map[h_str.lower()] = idx

    id_col = col_map.get(key_field.lower())
    if id_col is None:
        return header_row, col_map, header_row + 1

    data_start = header_row + 1
    for i in range(header_row + 1, min(header_row + 5, len(df))):
        cell = str(df.iloc[i, id_col]).strip().lower() if pd.notna(df.iloc[i, id_col]) else ""
        if cell in ("int", "string", "int[]", "int[][]", "string[]", "string[][]", "float"):
            data_start = i + 1
            continue
        if cell in ("key", "index", "key,set", "set", ""):
            data_start = i + 1
            continue
        first_val = df.iloc[i, id_col]
        if pd.notna(first_val):
            try:
                int(float(str(first_val)))
                data_start = i
                break
            except (ValueError, TypeError):
                data_start = i + 1

    return header_row, col_map, data_start


def parse_trigger_condition(raw_value) -> list[dict]:
    if pd.isna(raw_value):
        return []
    text = str(raw_value).strip()
    if not text or text == "nan":
        return []

    conditions = []
    groups = re.split(r'[;；]', text)
    for group in groups:
        group = group.strip().strip('()')
        if not group:
            continue
        parts = re.split(r'[,，]', group)
        if len(parts) >= 1:
            try:
                cond_id = int(float(parts[0].strip()))
                min_val = int(float(parts[1].strip())) if len(parts) > 1 else None
                max_val = int(float(parts[2].strip())) if len(parts) > 2 else None
                format_error = len(parts) < 2
                conditions.append({"id": cond_id, "min": min_val, "max": max_val, "formatError": format_error})
            except (ValueError, TypeError):
                conditions.append({"id": None, "min": None, "max": None, "formatError": True})
    return conditions


def read_conditions(folder_path: str) -> dict:
    file_path = os.path.join(folder_path, "Condition.xlsx")
    if not os.path.exists(file_path):
        return {}

    try:
        df = pd.read_excel(file_path, sheet_name="condition", header=None)
    except Exception:
        df = pd.read_excel(file_path, sheet_name=0, header=None)

    _, col_map, data_start = find_header_and_data_start(df, "Id")
    if col_map is None:
        return {}

    id_col = col_map.get("id")
    name_col = col_map.get("name")
    condition_col = col_map.get("condition")
    if id_col is None or name_col is None:
        return {}

    result = {}
    for i in range(data_start, len(df)):
        raw_id = df.iloc[i, id_col]
        if pd.isna(raw_id):
            continue
        try:
            cid = int(float(str(raw_id)))
        except (ValueError, TypeError):
            continue
        raw_name = df.iloc[i, name_col]
        name = str(raw_name).strip() if pd.notna(raw_name) else f"条件{cid}"
        if not name or name == "nan":
            name = f"条件{cid}"
        cond_type = None
        if condition_col is not None:
            raw_type = df.iloc[i, condition_col]
            if pd.notna(raw_type):
                try:
                    cond_type = int(float(str(raw_type)))
                except (ValueError, TypeError):
                    pass
        result[cid] = {"name": name, "condType": cond_type}

    return result


def read_story_groups(folder_path: str) -> list[dict]:
    file_path = os.path.join(folder_path, "StoryGroup.xlsx")
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"StoryGroup.xlsx not found in {folder_path}")

    df = pd.read_excel(file_path, sheet_name="Sheet1", header=None)

    _, col_map, data_start = find_header_and_data_start(df, "StoryGroupId")
    if col_map is None:
        raise ValueError("Cannot find header row with StoryGroupId column")

    id_col = col_map.get("storygroupid")
    name_col = col_map.get("name")
    trigger_col = col_map.get("triggercondition")
    selfadd_col = col_map.get("selfaddcondition")

    if id_col is None or name_col is None:
        raise ValueError("StoryGroupId or Name column not found")

    condition_map = read_conditions(folder_path)

    results = []
    for i in range(data_start, len(df)):
        raw_id = df.iloc[i, id_col]
        raw_name = df.iloc[i, name_col]

        if pd.isna(raw_id):
            continue

        try:
            story_id = int(float(str(raw_id)))
        except (ValueError, TypeError):
            continue

        name = str(raw_name).strip() if pd.notna(raw_name) else ""
        if not name or name == "nan":
            name = f"(unnamed-{story_id})"

        trigger_conditions = []
        if trigger_col is not None:
            raw_trigger = df.iloc[i, trigger_col]
            parsed = parse_trigger_condition(raw_trigger)
            for cond in parsed:
                if cond["id"] is None:
                    trigger_conditions.append({
                        "id": None,
                        "name": "格式错误",
                        "min": None,
                        "max": None,
                        "formatError": True
                    })
                else:
                    cond_info = condition_map.get(cond["id"], {"name": f"条件{cond['id']}", "condType": None})
                    trigger_conditions.append({
                        "id": cond["id"],
                        "name": cond_info["name"],
                        "min": cond["min"],
                        "max": cond["max"],
                        "formatError": cond["formatError"]
                    })

        self_add_conditions = []
        if selfadd_col is not None:
            raw_selfadd = df.iloc[i, selfadd_col]
            if pd.notna(raw_selfadd):
                text = str(raw_selfadd).strip()
                if text and text != "nan":
                    parts = re.split(r'[,，;；\s]+', text)
                    for p in parts:
                        p = p.strip().strip('()')
                        if p:
                            try:
                                sa_id = int(float(p))
                                cond_info = condition_map.get(sa_id, {"name": f"条件{sa_id}", "condType": None})
                                self_add_conditions.append({
                                    "id": sa_id,
                                    "name": cond_info["name"],
                                    "condType": cond_info["condType"]
                                })
                            except (ValueError, TypeError):
                                pass

        results.append({
            "id": story_id,
            "name": name,
            "conditions": trigger_conditions,
            "selfAddConditions": self_add_conditions
        })

    return results


def lookup_table(folder_path: str, filename: str, sheet_name, id_field: str) -> dict:
    file_path = os.path.join(folder_path, filename)
    if not os.path.exists(file_path):
        return {}

    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
    except Exception:
        df = pd.read_excel(file_path, sheet_name=0, header=None)

    _, col_map, data_start = find_header_and_data_start(df, id_field)
    if col_map is None:
        return {}

    id_col = col_map.get(id_field.lower())
    name_col = col_map.get("name")
    if id_col is None or name_col is None:
        return {}

    result = {}
    for i in range(data_start, len(df)):
        raw_id = df.iloc[i, id_col]
        if pd.isna(raw_id):
            continue
        try:
            rid = int(float(str(raw_id)))
        except (ValueError, TypeError):
            continue
        raw_name = df.iloc[i, name_col]
        name = str(raw_name).strip() if pd.notna(raw_name) else ""
        result[rid] = name

    return result


def read_condition_detail(folder_path: str, condition_id: int) -> dict | None:
    file_path = os.path.join(folder_path, "Condition.xlsx")
    if not os.path.exists(file_path):
        return None

    try:
        df = pd.read_excel(file_path, sheet_name="condition", header=None)
    except Exception:
        df = pd.read_excel(file_path, sheet_name=0, header=None)

    _, col_map, data_start = find_header_and_data_start(df, "Id")
    if col_map is None:
        return None

    id_col = col_map.get("id")
    name_col = col_map.get("name")
    condition_col = col_map.get("condition")
    params_col = col_map.get("conditionparams")
    datatype_col = col_map.get("datatype")

    if id_col is None:
        return None

    target_row = None
    for i in range(data_start, len(df)):
        raw_id = df.iloc[i, id_col]
        if pd.isna(raw_id):
            continue
        try:
            cid = int(float(str(raw_id)))
        except (ValueError, TypeError):
            continue
        if cid == condition_id:
            target_row = i
            break

    if target_row is None:
        return None

    def get_int(col_idx):
        if col_idx is None:
            return None
        val = df.iloc[target_row, col_idx]
        if pd.isna(val):
            return None
        try:
            return int(float(str(val)))
        except (ValueError, TypeError):
            return None

    def get_str(col_idx):
        if col_idx is None:
            return ""
        val = df.iloc[target_row, col_idx]
        if pd.isna(val):
            return ""
        return str(val).strip()

    def get_int_list(col_idx):
        if col_idx is None:
            return []
        val = df.iloc[target_row, col_idx]
        if pd.isna(val):
            return []
        text = str(val).strip()
        if not text or text == "nan":
            return []
        parts = re.split(r'[,，;；\s]+', text)
        result = []
        for p in parts:
            p = p.strip().strip('()')
            if p:
                try:
                    result.append(int(float(p)))
                except (ValueError, TypeError):
                    pass
        return result

    cond_id = condition_id
    cond_name = get_str(name_col) or f"条件{cond_id}"
    cond_type = get_int(condition_col)
    cond_params = get_int_list(params_col)
    data_type = get_int(datatype_col)

    detail = {
        "id": cond_id,
        "name": cond_name,
        "type": cond_type,
        "params": cond_params,
        "dataType": data_type,
        "refs": []
    }

    if cond_type == 7 and cond_params:
        task_map = lookup_table(folder_path, "Task.xlsx", "Sheet1", "Id")
        for pid in cond_params:
            detail["refs"].append({
                "table": "Task",
                "id": pid,
                "name": task_map.get(pid, f"任务{pid}")
            })

    elif cond_type == 8 and cond_params:
        equip_map = lookup_table(folder_path, "EquipBase.xlsx", "装备基底", "ID")
        for pid in cond_params:
            detail["refs"].append({
                "table": "EquipBase",
                "id": pid,
                "name": equip_map.get(pid, f"装备{pid}")
            })

    elif cond_type == 9 and cond_params:
        item_map = lookup_table(folder_path, "Item.xlsx", "item", "Id")
        for pid in cond_params:
            detail["refs"].append({
                "table": "Item",
                "id": pid,
                "name": item_map.get(pid, f"道具{pid}")
            })

    elif cond_type == 10 and cond_params:
        stage_map = lookup_table(folder_path, "Stage.xlsx", "stage", "StageId")
        if len(cond_params) >= 3 and cond_params[0] == 12:
            low = cond_params[1]
            high = cond_params[2]
            for sid in range(low, high + 1):
                detail["refs"].append({
                    "table": "Stage",
                    "id": sid,
                    "name": stage_map.get(sid, f"关卡{sid}")
                })
        else:
            for pid in cond_params:
                detail["refs"].append({
                    "table": "Stage",
                    "id": pid,
                    "name": stage_map.get(pid, f"关卡{pid}")
                })

    elif cond_type == 13 and cond_params:
        cond_map = read_conditions(folder_path)
        for pid in cond_params[:2]:
            cond_info = cond_map.get(pid, {"name": f"条件{pid}", "condType": None})
            detail["refs"].append({
                "table": "Condition",
                "id": pid,
                "name": cond_info["name"]
            })

    elif cond_type == 14 and cond_params:
        talent_map = lookup_table(folder_path, "Talent.xlsx", "Sheet1", "TalentId")
        for pid in cond_params:
            detail["refs"].append({
                "table": "Talent",
                "id": pid,
                "name": talent_map.get(pid, f"天赋{pid}")
            })

    return detail


def lookup_table_multi(folder_path: str, filename: str, sheet_name, id_field: str, fields: list[str]) -> dict:
    file_path = os.path.join(folder_path, filename)
    if not os.path.exists(file_path):
        return {}

    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
    except Exception:
        df = pd.read_excel(file_path, sheet_name=0, header=None)

    _, col_map, data_start = find_header_and_data_start(df, id_field)
    if col_map is None:
        return {}

    id_col = col_map.get(id_field.lower())
    if id_col is None:
        return {}

    field_cols = {}
    for f in fields:
        fc = col_map.get(f.lower())
        if fc is not None:
            field_cols[f.lower()] = fc

    result = {}
    for i in range(data_start, len(df)):
        raw_id = df.iloc[i, id_col]
        if pd.isna(raw_id):
            continue
        try:
            rid = int(float(str(raw_id)))
        except (ValueError, TypeError):
            continue
        row_data = {}
        for f_lower, fc in field_cols.items():
            val = df.iloc[i, fc]
            if pd.isna(val):
                row_data[f_lower] = None
            else:
                row_data[f_lower] = str(val).strip()
        result[rid] = row_data

    return result


def lookup_map_events(folder_path: str) -> tuple[dict, dict]:
    file_path = os.path.join(folder_path, "MapEvent.xlsx")
    if not os.path.exists(file_path):
        return {}, {}

    try:
        df = pd.read_excel(file_path, sheet_name="Sheet1", header=None)
    except Exception:
        df = pd.read_excel(file_path, sheet_name=0, header=None)

    _, col_map, data_start = find_header_and_data_start(df, "EventId")
    if col_map is None:
        return {}, {}

    id_col = col_map.get("eventid")
    desc_col = col_map.get("desc1")
    mapid_col = col_map.get("mapid")
    pos_col = col_map.get("position")

    by_id = {}
    by_pos = {}

    for i in range(data_start, len(df)):
        raw_id = df.iloc[i, id_col] if id_col is not None else None
        if raw_id is not None and not pd.isna(raw_id):
            try:
                eid = int(float(str(raw_id)))
            except (ValueError, TypeError):
                continue
        else:
            continue

        desc = ""
        if desc_col is not None:
            raw_desc = df.iloc[i, desc_col]
            desc = str(raw_desc).strip() if pd.notna(raw_desc) else ""

        map_id = None
        if mapid_col is not None:
            raw_map = df.iloc[i, mapid_col]
            if pd.notna(raw_map):
                try:
                    map_id = int(float(str(raw_map)))
                except (ValueError, TypeError):
                    pass

        position = ""
        if pos_col is not None:
            raw_pos = df.iloc[i, pos_col]
            position = str(raw_pos).strip() if pd.notna(raw_pos) else ""

        by_id[eid] = {"id": eid, "desc": desc, "mapId": map_id, "position": position, "eventId": eid}

        if map_id is not None and position:
            norm_pos = ",".join([str(int(float(p.strip()))) for p in position.split(",") if p.strip()])
            key = f"{map_id},{norm_pos}"
            if key not in by_pos:
                by_pos[key] = {"id": eid, "desc": desc, "eventId": eid}

    return by_id, by_pos


def read_task_detail(folder_path: str, task_id: int) -> dict | None:
    file_path = os.path.join(folder_path, "Task.xlsx")
    if not os.path.exists(file_path):
        return None

    try:
        df = pd.read_excel(file_path, sheet_name="Sheet1", header=None)
    except Exception:
        df = pd.read_excel(file_path, sheet_name=0, header=None)

    _, col_map, data_start = find_header_and_data_start(df, "Id")
    if col_map is None:
        return None

    id_col = col_map.get("id")
    if id_col is None:
        return None

    target_row = None
    all_ids = []
    for i in range(data_start, len(df)):
        raw_id = df.iloc[i, id_col]
        if pd.isna(raw_id):
            continue
        try:
            tid = int(float(str(raw_id)))
        except (ValueError, TypeError):
            continue
        all_ids.append(tid)
        if tid == task_id:
            target_row = i

    if target_row is None:
        return None

    duplicate_ids = [x for x in all_ids if all_ids.count(x) > 1]

    def get_val(field_name):
        fc = col_map.get(field_name.lower())
        if fc is None:
            return None
        val = df.iloc[target_row, fc]
        if pd.isna(val):
            return None
        return str(val).strip()

    def get_int_val(field_name):
        v = get_val(field_name)
        if v is None or v == "" or v == "nan":
            return None
        try:
            return int(float(v))
        except (ValueError, TypeError):
            return None

    condition_map = read_conditions(folder_path)
    item_map = lookup_table(folder_path, "Item.xlsx", "item", "Id")
    task_map = lookup_table(folder_path, "Task.xlsx", "Sheet1", "Id")
    story_map = lookup_table(folder_path, "Story.xlsx", "Sheet1", "Id")
    task_group_map = lookup_table(folder_path, "TaskGroup.xlsx", "Sheet1", "TaskGroupId")
    realm_level_map = lookup_table(folder_path, "RealmL_Level.xlsx", "Sheet1", "Level")
    map_events_by_id, map_events_by_pos = lookup_map_events(folder_path)

    def parse_condition_field(raw):
        if raw is None or raw == "" or raw == "nan":
            return []
        return parse_trigger_condition(raw)

    def resolve_conditions(raw):
        parsed = parse_condition_field(raw)
        result = []
        for c in parsed:
            if c["id"] is None:
                result.append({"id": None, "name": "格式错误", "min": None, "max": None, "formatError": True})
            else:
                cond_info = condition_map.get(c["id"], {"name": f"条件{c['id']}", "condType": None})
                result.append({"id": c["id"], "name": cond_info["name"], "min": c["min"], "max": c["max"], "formatError": c["formatError"]})
        return result

    def parse_trigger_field(raw):
        if raw is None or raw == "" or raw == "nan":
            return []
        groups = raw.split(";")
        result = []
        for g in groups:
            g = g.strip()
            if not g:
                continue
            parts = g.split(",")
            if not parts:
                continue
            try:
                trigger_type = int(float(parts[0].strip()))
            except (ValueError, TypeError):
                continue

            if trigger_type == 1:
                item_id = int(float(parts[1].strip())) if len(parts) > 1 else None
                count = int(float(parts[2].strip())) if len(parts) > 2 else 0
                item_name = item_map.get(item_id, f"道具{item_id}") if item_id else "未知"
                result.append({"type": 1, "itemId": item_id, "itemName": item_name, "count": count})

            elif trigger_type == 2:
                story_id = int(float(parts[1].strip())) if len(parts) > 1 else None
                story_name = story_map.get(story_id, f"剧情{story_id}") if story_id else "未知"
                result.append({"type": 2, "storyId": story_id, "storyName": story_name, "isStoryTriggerPoint": True})

            elif trigger_type == 3:
                event_id = int(float(parts[1].strip())) if len(parts) > 1 else None
                event_range = parts[2].strip() if len(parts) > 2 else "0"
                event_info = map_events_by_id.get(event_id, {})
                event_name = event_info.get("desc", f"事件{event_id}") if event_id else "未知"
                result.append({"type": 3, "eventId": event_id, "eventName": event_name, "eventRange": event_range})

            elif trigger_type == 4:
                modify_val = parts[1].strip() if len(parts) > 1 else "0"
                cond_ids = []
                for p in parts[2:]:
                    try:
                        cond_ids.append(int(float(p.strip())))
                    except (ValueError, TypeError):
                        pass
                conditions = []
                for cid in cond_ids:
                    cond_info = condition_map.get(cid, {"name": f"条件{cid}", "condType": None})
                    conditions.append({"id": cid, "name": cond_info["name"]})
                result.append({"type": 4, "condNum": modify_val, "conditions": conditions})

            else:
                result.append({"type": trigger_type, "raw": g})

        return result

    def parse_reward_field(raw):
        if raw is None or raw == "" or raw == "nan":
            return []
        groups = raw.split(";")
        result = []
        for g in groups:
            g = g.strip()
            if not g:
                continue
            parts = g.split(",")
            if len(parts) >= 2:
                try:
                    item_id = int(float(parts[0].strip()))
                    count = int(float(parts[1].strip()))
                    item_name = item_map.get(item_id, f"道具{item_id}")
                    result.append({"itemId": item_id, "name": item_name, "count": count})
                except (ValueError, TypeError):
                    pass
        return result

    def parse_remove_item_field(raw):
        if raw is None or raw == "" or raw == "nan":
            return []
        parts = raw.split(",")
        if len(parts) >= 2:
            try:
                item_id = int(float(parts[0].strip()))
                count = int(float(parts[1].strip()))
                item_name = item_map.get(item_id, f"道具{item_id}")
                return [{"itemId": item_id, "name": item_name, "count": count}]
            except (ValueError, TypeError):
                pass
        return []

    def parse_self_add_condition_field(raw):
        if raw is None or raw == "" or raw == "nan":
            return []
        groups = raw.split(";")
        result = []
        for g in groups:
            g = g.strip()
            if not g:
                continue
            parts = g.split(",")
            if len(parts) >= 2:
                try:
                    cid = int(float(parts[0].strip()))
                    val = int(float(parts[1].strip()))
                    cond_info = condition_map.get(cid, {"name": f"条件{cid}", "condType": None})
                    result.append({"id": cid, "name": cond_info["name"], "value": val})
                except (ValueError, TypeError):
                    pass
        return result

    def parse_track_field(raw):
        if raw is None or raw == "" or raw == "nan" or raw == "0":
            return None
        parts = raw.split(",")
        if len(parts) >= 3:
            try:
                map_id = int(float(parts[0].strip()))
                x = int(float(parts[1].strip()))
                y = int(float(parts[2].strip()))
                pos_key = f"{map_id},{x},{y}"
                event_info = map_events_by_pos.get(pos_key)
                if event_info:
                    return {
                        "found": True,
                        "eventId": event_info["eventId"],
                        "eventName": event_info["desc"],
                        "mapId": map_id,
                        "x": x,
                        "y": y
                    }
                else:
                    return {"found": False, "mapId": map_id, "x": x, "y": y}
            except (ValueError, TypeError):
                return {"found": False, "raw": raw}
        return None

    task_type = get_int_val("Type")
    sub_type = get_int_val("SubType")
    group_raw = get_val("Group")
    group_info = None
    if task_type in (3, 4) and group_raw and group_raw != "nan":
        parts = group_raw.split(",")
        if len(parts) >= 2:
            try:
                gid = int(float(parts[0].strip()))
                weight = int(float(parts[1].strip()))
                gname = task_group_map.get(gid, f"分组{gid}")
                group_info = {"id": gid, "name": gname, "weight": weight}
            except (ValueError, TypeError):
                group_info = {"id": None, "name": "格式错误", "weight": None}

    realm_level_raw = get_int_val("RealmLevel")
    realm_level_info = None
    if task_type in (3, 4) and realm_level_raw is not None:
        rname = realm_level_map.get(realm_level_raw, f"等级{realm_level_raw}")
        realm_level_info = {"level": realm_level_raw, "name": rname}

    next_task_raw = get_int_val("NextTask")
    next_task_info = None
    if next_task_raw is not None:
        next_task_name = task_map.get(next_task_raw, f"任务{next_task_raw}")
        next_task_info = {"id": next_task_raw, "name": next_task_name}

    result = {
        "id": task_id,
        "name": get_val("Name") or f"任务{task_id}",
        "dsc": get_val("Dsc") or "",
        "type": task_type,
        "subType": sub_type,
        "group": group_info,
        "realmLevel": realm_level_info,
        "difficulty": get_int_val("Difficulty"),
        "countdown": get_int_val("Countdown"),
        "acceptCondition": resolve_conditions(get_val("AcceptCondition")),
        "isRepeat": get_int_val("IsRepeat"),
        "isHide": get_int_val("IsHide"),
        "acceptTrigger": parse_trigger_field(get_val("AcceptTrigger")),
        "allowDirectReturn": get_int_val("AllowDirectReturn"),
        "deleteTrigger": parse_trigger_field(get_val("DeleteTrigger")),
        "failTrigger": parse_trigger_field(get_val("FailTrigger")),
        "completeCondition": resolve_conditions(get_val("CompleteCondition")),
        "removeItem": parse_remove_item_field(get_val("RemoveItem")),
        "reward": parse_reward_field(get_val("Reward")),
        "addFavor": get_int_val("AddFavor"),
        "nextTask": next_task_info,
        "selfAddCondition": parse_self_add_condition_field(get_val("SelfAddCondition")),
        "inProgressTrack": parse_track_field(get_val("InProgressTrack")),
        "completeTrack": parse_track_field(get_val("CompleteTrack")),
        "isDuplicate": task_id in duplicate_ids,
        "taskType": task_type
    }

    return result


def get_task_group_options(folder_path: str) -> list:
    tg_map = lookup_table(folder_path, "TaskGroup.xlsx", "Sheet1", "TaskGroupId")
    return [{"id": k, "name": v} for k, v in tg_map.items()]


def get_realm_level_options(folder_path: str) -> list:
    rl_map = lookup_table(folder_path, "RealmL_Level.xlsx", "Sheet1", "Level")
    return [{"id": k, "name": v} for k, v in rl_map.items()]


@app.get("/api/current-folder")
async def get_current_folder():
    return {"path": DATA_DIR}


@app.post("/api/set-folder")
async def set_folder(req: FolderRequest):
    global DATA_DIR
    folder = req.path.strip()
    if not os.path.isdir(folder):
        raise HTTPException(status_code=400, detail=f"Folder not found: {folder}")
    story_file = os.path.join(folder, "StoryGroup.xlsx")
    if not os.path.exists(story_file):
        raise HTTPException(status_code=400, detail="StoryGroup.xlsx not found in the specified folder")
    DATA_DIR = folder
    try:
        with open(_CONFIG_FILE, "w", encoding="utf-8") as f:
            f.write(folder)
    except Exception:
        pass
    return {"status": "ok", "path": DATA_DIR}


@app.get("/api/current-asset-folder")
async def get_current_asset_folder():
    return {"path": ASSET_DIR}


@app.post("/api/set-asset-folder")
async def set_asset_folder(req: FolderRequest):
    global ASSET_DIR
    folder = req.path.strip()
    if not os.path.isdir(folder):
        raise HTTPException(status_code=400, detail=f"Folder not found: {folder}")
    ASSET_DIR = folder
    try:
        with open(_ASSET_CONFIG_FILE, "w", encoding="utf-8") as f:
            f.write(folder)
    except Exception:
        pass
    return {"status": "ok", "path": ASSET_DIR}


@app.get("/api/story-groups")
async def get_story_groups():
    if DATA_DIR is None:
        raise HTTPException(status_code=400, detail="No folder configured. Use /api/set-folder first.")
    try:
        groups = read_story_groups(DATA_DIR)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    return {"groups": groups}


@app.get("/api/condition/{condition_id}")
async def get_condition_detail(condition_id: int):
    if DATA_DIR is None:
        raise HTTPException(status_code=400, detail="No folder configured.")
    detail = read_condition_detail(DATA_DIR, condition_id)
    if detail is None:
        raise HTTPException(status_code=404, detail=f"Condition {condition_id} not found")
    return detail


@app.get("/api/task/{task_id}")
async def get_task_detail(task_id: int):
    if DATA_DIR is None:
        raise HTTPException(status_code=400, detail="No folder configured.")
    detail = read_task_detail(DATA_DIR, task_id)
    if detail is None:
        raise HTTPException(status_code=404, detail=f"Task {task_id} not found")
    return detail


@app.get("/api/task-list")
async def get_task_list():
    if DATA_DIR is None:
        raise HTTPException(status_code=400, detail="No folder configured.")
    file_path = os.path.join(DATA_DIR, "Task.xlsx")
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Task.xlsx not found")
    df = pd.read_excel(file_path, sheet_name="Sheet1", header=None)
    _, col_map, data_start = find_header_and_data_start(df, "Id")
    if col_map is None:
        return []
    id_col = col_map.get("id")
    name_col = col_map.get("name")
    if id_col is None:
        return []
    result = []
    for i in range(data_start, len(df)):
        raw_id = df.iloc[i, id_col]
        if pd.isna(raw_id):
            continue
        try:
            tid = int(float(str(raw_id)))
        except (ValueError, TypeError):
            continue
        name = ""
        if name_col is not None:
            raw_name = df.iloc[i, name_col]
            if pd.notna(raw_name):
                name = str(raw_name).strip()
                if name == "nan":
                    name = ""
        result.append({"id": tid, "name": name or f"任务{tid}"})
    return result


@app.get("/api/task-group-options")
async def get_task_group_opts():
    if DATA_DIR is None:
        raise HTTPException(status_code=400, detail="No folder configured.")
    return get_task_group_options(DATA_DIR)


@app.get("/api/realm-level-options")
async def get_realm_level_opts():
    if DATA_DIR is None:
        raise HTTPException(status_code=400, detail="No folder configured.")
    return get_realm_level_options(DATA_DIR)


@app.get("/api/field-defaults")
async def get_field_defaults():
    if DATA_DIR is None:
        raise HTTPException(status_code=400, detail="No folder configured.")
    item_map = lookup_table(DATA_DIR, "Item.xlsx", "item", "Id")
    condition_map = read_conditions(DATA_DIR)
    story_map = lookup_table(DATA_DIR, "Story.xlsx", "Sheet1", "Id")
    map_events_by_id, _ = lookup_map_events(DATA_DIR)
    first_item_id = None
    first_item_name = ""
    if item_map:
        first_item_id = min(item_map.keys())
        first_item_name = item_map[first_item_id]
    first_cond_id = None
    first_cond_name = ""
    if condition_map:
        first_cond_id = min(condition_map.keys())
        first_cond_name = condition_map[first_cond_id]["name"]
    first_story_id = None
    first_story_name = ""
    if story_map:
        first_story_id = min(story_map.keys())
        first_story_name = story_map[first_story_id]
    first_event_id = None
    first_event_name = ""
    if map_events_by_id:
        first_event_id = min(map_events_by_id.keys())
        first_event_name = map_events_by_id[first_event_id].get("desc", "")
    return {
        "item": {"id": first_item_id, "name": first_item_name},
        "condition": {"id": first_cond_id, "name": first_cond_name},
        "story": {"id": first_story_id, "name": first_story_name},
        "event": {"id": first_event_id, "name": first_event_name},
    }


class ExportRequest(BaseModel):
    changes: Dict[str, Any]


FIELD_TO_COL = {
    "id": "Id", "name": "Name", "dsc": "Dsc", "type": "Type", "subType": "SubType",
    "group": "Group", "realmLevel": "RealmLevel", "difficulty": "Difficulty",
    "countdown": "Countdown", "isRepeat": "IsRepeat", "isHide": "IsHide",
    "allowDirectReturn": "AllowDirectReturn", "addFavor": "AddFavor", "nextTask": "NextTask",
    "groupWeight": "Group",
    "desc": "Desc", "gender": "Gender", "pos": "Pos", "spine": "Spine", "icon": "Icon",
}


@app.get("/api/lookup-name")
async def lookup_name(table: str, id: int):
    if DATA_DIR is None:
        raise HTTPException(status_code=400, detail="No folder configured.")
    name = None
    if table == "Item":
        item_map = lookup_table(DATA_DIR, "Item.xlsx", "item", "Id")
        name = item_map.get(id)
    elif table == "Condition":
        cond_map = read_conditions(DATA_DIR)
        cond = cond_map.get(id)
        if cond:
            name = cond["name"]
    elif table == "Story":
        story_map = lookup_table(DATA_DIR, "Story.xlsx", "Sheet1", "Id")
        name = story_map.get(id)
    elif table == "MapEvent":
        map_events_by_id, _ = lookup_map_events(DATA_DIR)
        ev = map_events_by_id.get(id)
        if ev:
            name = ev.get("desc", "")
    elif table == "Task":
        task_map = lookup_table(DATA_DIR, "Task.xlsx", "Sheet1", "Id")
        name = task_map.get(id)
    if name is None:
        return {"found": False}
    return {"found": True, "name": name}


@app.post("/api/export")
async def export_changes(req: ExportRequest):
    if DATA_DIR is None:
        raise HTTPException(status_code=400, detail="No folder configured.")

    export_base = os.path.join(os.path.dirname(__file__), "导出文件")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    export_dir = os.path.join(export_base, timestamp)
    os.makedirs(export_dir, exist_ok=True)

    files_to_modify = {}
    for key, item in req.changes.items():
        file_name = item.get("file", "Task.xlsx")
        if file_name not in files_to_modify:
            files_to_modify[file_name] = []
        files_to_modify[file_name].append(item)

    exported_files = []
    for file_name, items in files_to_modify.items():
        src_path = os.path.join(DATA_DIR, file_name)
        if not os.path.exists(src_path):
            continue
        dst_path = os.path.join(export_dir, file_name)
        shutil.copy2(src_path, dst_path)

        wb = load_workbook(dst_path)
        if file_name == "EquipBase.xlsx":
            ws = wb["装备基底"] if "装备基底" in wb.sheetnames else wb.active
        else:
            ws = wb.active

        header_row_idx = None
        col_map = {}
        for row_idx in range(1, min(11, ws.max_row + 1)):
            for col_idx in range(1, ws.max_column + 1):
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val and str(cell_val).strip() in ("Id", "ID", "EventId"):
                    header_row_idx = row_idx
                    break
            if header_row_idx:
                break

        if header_row_idx is None:
            continue

        for col_idx in range(1, ws.max_column + 1):
            cell_val = ws.cell(row=header_row_idx, column=col_idx).value
            if cell_val:
                col_map[str(cell_val).strip()] = col_idx

        data_start = header_row_idx + 1
        for row_idx in range(header_row_idx + 1, min(header_row_idx + 5, ws.max_row + 1)):
            id_col_idx = col_map.get("Id") or col_map.get("ID") or col_map.get("EventId")
            if id_col_idx:
                cell_val = str(ws.cell(row=row_idx, column=id_col_idx).value or "").strip().lower()
                if cell_val in ("int", "string", "int[]", "int[][]", "string[]", "string[][]", "float", "key", "index", "key,set", "set", ""):
                    data_start = row_idx + 1
                else:
                    break

        id_col_idx = col_map.get("Id") or col_map.get("ID") or col_map.get("EventId")
        for item in items:
            task_id = item.get("taskId")
            is_delete = item.get("_delete", False)
            is_new = item.get("_new", False)
            changes = item.get("changes", {})
            if not task_id:
                continue
            if not is_delete and not is_new and not changes:
                continue

            if is_new:
                new_row = ws.max_row + 1
                ws.cell(row=new_row, column=id_col_idx).value = task_id
                for field, value in changes.items():
                    if field in ("acceptCondition", "completeCondition", "acceptTrigger", "deleteTrigger", "failTrigger", "removeItem", "reward", "selfAddCondition", "inProgressTrack", "completeTrack"):
                        col_map_name = {
                            "acceptCondition": "AcceptCondition", "completeCondition": "CompleteCondition",
                            "acceptTrigger": "AcceptTrigger", "deleteTrigger": "DeleteTrigger",
                            "failTrigger": "FailTrigger", "removeItem": "RemoveItem",
                            "reward": "Reward", "selfAddCondition": "SelfAddCondition",
                            "inProgressTrack": "InProgressTrack", "completeTrack": "CompleteTrack",
                        }.get(field)
                        if col_map_name and col_map_name in col_map:
                            ws.cell(row=new_row, column=col_map[col_map_name]).value = None if value == '' else value
                    else:
                        col_name = FIELD_TO_COL.get(field)
                        if col_name:
                            actual_col = col_map.get(col_name) or col_map.get(col_name.upper())
                            if actual_col:
                                ws.cell(row=new_row, column=actual_col).value = value
                continue

            target_row = None
            for row_idx in range(data_start, ws.max_row + 1):
                cell_val = ws.cell(row=row_idx, column=id_col_idx).value
                if cell_val is not None:
                    try:
                        if int(float(str(cell_val))) == task_id:
                            target_row = row_idx
                            break
                    except (ValueError, TypeError):
                        continue

            if target_row is None:
                continue

            if is_delete:
                ws.delete_rows(target_row)
                continue

            for field, value in changes.items():
                if field in ("acceptCondition", "completeCondition", "acceptTrigger", "deleteTrigger", "failTrigger", "removeItem", "reward", "selfAddCondition", "inProgressTrack", "completeTrack"):
                    col_map_name = {
                        "acceptCondition": "AcceptCondition", "completeCondition": "CompleteCondition",
                        "acceptTrigger": "AcceptTrigger", "deleteTrigger": "DeleteTrigger",
                        "failTrigger": "FailTrigger", "removeItem": "RemoveItem",
                        "reward": "Reward", "selfAddCondition": "SelfAddCondition",
                        "inProgressTrack": "InProgressTrack", "completeTrack": "CompleteTrack",
                    }.get(field)
                    if col_map_name:
                        col_idx = col_map.get(col_map_name)
                        if col_idx:
                            ws.cell(row=target_row, column=col_idx).value = None if value == '' else value
                elif field.startswith("acceptCond_") or field.startswith("completeCond_"):
                    parts = field.split("_")
                    prefix = parts[0]
                    idx = int(parts[1])
                    min_or_max = parts[2]
                    col_name = "AcceptCondition" if prefix == "acceptCond" else "CompleteCondition"
                    col_idx = col_map.get(col_name)
                    if col_idx:
                        cell_val = ws.cell(row=target_row, column=col_idx).value
                        if cell_val:
                            cond_list = str(cell_val).split("|")
                            if idx < len(cond_list):
                                cond_parts = cond_list[idx].split(",")
                                if min_or_max == "min" and len(cond_parts) >= 2:
                                    cond_parts[1] = str(value)
                                elif min_or_max == "max" and len(cond_parts) >= 3:
                                    cond_parts[2] = str(value)
                                cond_list[idx] = ",".join(cond_parts)
                                ws.cell(row=target_row, column=col_idx).value = "|".join(cond_list)
                elif field.startswith("removeItem_"):
                    parts = field.split("_")
                    idx = int(parts[1])
                    col_idx = col_map.get("RemoveItem")
                    if col_idx:
                        cell_val = ws.cell(row=target_row, column=col_idx).value
                        if cell_val:
                            item_list = str(cell_val).split("|")
                            if idx < len(item_list):
                                item_parts = item_list[idx].split(",")
                                if len(item_parts) >= 2:
                                    item_parts[1] = str(value)
                                item_list[idx] = ",".join(item_parts)
                                ws.cell(row=target_row, column=col_idx).value = "|".join(item_list)
                elif field.startswith("reward_"):
                    parts = field.split("_")
                    idx = int(parts[1])
                    col_idx = col_map.get("Reward")
                    if col_idx:
                        cell_val = ws.cell(row=target_row, column=col_idx).value
                        if cell_val:
                            reward_list = str(cell_val).split("|")
                            if idx < len(reward_list):
                                reward_parts = reward_list[idx].split(",")
                                if len(reward_parts) >= 2:
                                    reward_parts[1] = str(value)
                                reward_list[idx] = ",".join(reward_parts)
                                ws.cell(row=target_row, column=col_idx).value = "|".join(reward_list)
                elif field == "groupWeight":
                    col_idx = col_map.get("Group")
                    if col_idx:
                        cell_val = ws.cell(row=target_row, column=col_idx).value
                        if cell_val:
                            gparts = str(cell_val).split(",")
                            if len(gparts) >= 2:
                                gparts[1] = str(value)
                            else:
                                gparts.append(str(value))
                            ws.cell(row=target_row, column=col_idx).value = ",".join(gparts)
                elif field == "group" and item.get("table") != "EquipBase":
                    col_idx = col_map.get("Group")
                    if col_idx:
                        old_val = ws.cell(row=target_row, column=col_idx).value
                        if old_val:
                            gparts = str(old_val).split(",")
                            gparts[0] = str(value)
                            ws.cell(row=target_row, column=col_idx).value = ",".join(gparts)
                        else:
                            ws.cell(row=target_row, column=col_idx).value = str(value)
                else:
                    col_name = FIELD_TO_COL.get(field)
                    if col_name:
                        col_idx = col_map.get(col_name) or col_map.get(col_name.upper())
                        if col_idx:
                            ws.cell(row=target_row, column=col_idx).value = value

        wb.save(dst_path)
        exported_files.append(file_name)

    return {"message": f"已导出 {len(exported_files)} 个文件到 app/导出文件/{timestamp}", "files": exported_files}


EQUIP_FIELDS = ["ID", "Name", "Desc", "Group", "Gender", "Pos", "Type", "Spine", "Icon"]


def read_equip_all():
    if DATA_DIR is None:
        return []
    file_path = os.path.join(DATA_DIR, "EquipBase.xlsx")
    if not os.path.exists(file_path):
        return []
    try:
        df = pd.read_excel(file_path, sheet_name="装备基底", header=None)
    except Exception:
        df = pd.read_excel(file_path, sheet_name=0, header=None)
    _, col_map, data_start = find_header_and_data_start(df, "ID")
    if col_map is None:
        return []
    id_col = col_map.get("id")
    if id_col is None:
        return []
    field_cols = {}
    for f in EQUIP_FIELDS:
        fc = col_map.get(f.lower())
        if fc is not None:
            field_cols[f] = fc
    result = []
    for i in range(data_start, len(df)):
        raw_id = df.iloc[i, id_col]
        if pd.isna(raw_id):
            continue
        try:
            rid = int(float(str(raw_id)))
        except (ValueError, TypeError):
            continue
        row = {"id": rid}
        for f, fc in field_cols.items():
            if f == "ID":
                continue
            val = df.iloc[i, fc]
            if pd.isna(val):
                row[f.lower()] = None
            else:
                try:
                    row[f.lower()] = int(float(str(val)))
                except (ValueError, TypeError):
                    row[f.lower()] = str(val).strip()
        result.append(row)
    return result


@app.get("/api/equip-list")
async def get_equip_list():
    if DATA_DIR is None:
        raise HTTPException(status_code=400, detail="No folder configured.")
    items = read_equip_all()
    return [{"id": e["id"], "name": e.get("name") or f"装备{e['id']}"} for e in items]


@app.get("/api/equip/{equip_id}")
async def get_equip_detail(equip_id: int):
    if DATA_DIR is None:
        raise HTTPException(status_code=400, detail="No folder configured.")
    items = read_equip_all()
    for e in items:
        if e["id"] == equip_id:
            return e
    raise HTTPException(status_code=404, detail=f"Equip {equip_id} not found")


@app.get("/api/spine-list")
async def get_spine_list():
    if ASSET_DIR is None:
        raise HTTPException(status_code=400, detail="No asset folder configured.")
    spine_dir = os.path.join(ASSET_DIR, "Assets", "Game", "Bundle", "Spine", "Hero")
    if not os.path.isdir(spine_dir):
        return []
    result = []
    gameres_spine = os.path.join(ASSET_DIR, "Assets", "Game", "GameRes", "Spine")
    spine_data_map = {}
    folder_map = {}
    if os.path.isdir(gameres_spine):
        for root, dirs, files in os.walk(gameres_spine):
            for f in files:
                if f.lower().endswith(".atlas.txt"):
                    name_no_ext = f[:-len(".atlas.txt")]
                    rel_path = os.path.relpath(root, os.path.join(ASSET_DIR, "Assets", "Game", "GameRes")).replace("\\", "/")
                    atlas_url = f"/api/gameres-file/{rel_path}/{f}"
                    skel_file = f"{name_no_ext}.skel.bytes"
                    png_file = f"{name_no_ext}.png"
                    skel_url = f"/api/gameres-file/{rel_path}/{skel_file}" if skel_file in files else None
                    png_url = f"/api/gameres-file/{rel_path}/{png_file}" if png_file in files else None
                    if name_no_ext not in spine_data_map:
                        spine_data_map[name_no_ext] = {
                            "atlas": atlas_url,
                            "skel": skel_url,
                            "image": png_url,
                            "preview": png_url
                        }
                        parts = rel_path.split("/")
                        folder_name = parts[1] if len(parts) > 1 else "Other"
                        folder_map[name_no_ext] = folder_name
    for name in os.listdir(spine_dir):
        if name.lower().endswith(".prefab"):
            prefab_name = name[:-len(".prefab")]
            data = spine_data_map.get(prefab_name, {})
            folder = folder_map.get(prefab_name, "Other")
            result.append({
                "name": prefab_name,
                "preview": data.get("preview"),
                "folder": folder,
                "atlas": data.get("atlas"),
                "skel": data.get("skel"),
                "image": data.get("image")
            })
    result.sort(key=lambda x: x["name"])
    return result


@app.get("/api/texture-list")
async def get_texture_list():
    if ASSET_DIR is None:
        raise HTTPException(status_code=400, detail="No asset folder configured.")
    tex_dir = os.path.join(ASSET_DIR, "Assets", "Game", "Bundle", "UI", "Texture")
    if not os.path.isdir(tex_dir):
        return []
    result = []
    for root, dirs, files in os.walk(tex_dir):
        for name in files:
            if name.lower().endswith((".png", ".jpg", ".jpeg", ".tga")) and not name.endswith(".meta"):
                rel_path = os.path.relpath(os.path.join(root, name), os.path.join(ASSET_DIR, "Assets", "Game", "Bundle")).replace("\\", "/")
                folder_rel = os.path.relpath(root, tex_dir).replace("\\", "/")
                if folder_rel == ".":
                    folder_rel = "Root"
                display_name = os.path.splitext(name)[0]
                result.append({"name": display_name, "file": name, "preview": f"/api/asset-file/{rel_path}", "folder": folder_rel})
    result.sort(key=lambda x: x["name"])
    return result


@app.get("/api/gameres-file/{path:path}")
async def get_gameres_file(path: str):
    if ASSET_DIR is None:
        raise HTTPException(status_code=400, detail="No asset folder configured.")
    full_path = os.path.join(ASSET_DIR, "Assets", "Game", "GameRes", path)
    if not os.path.isfile(full_path):
        raise HTTPException(status_code=404, detail="File not found")
    return FileResponse(full_path)


@app.get("/api/asset-file/{path:path}")
async def get_asset_file(path: str):
    if ASSET_DIR is None:
        raise HTTPException(status_code=400, detail="No asset folder configured.")
    full_path = os.path.join(ASSET_DIR, "Assets", "Game", "Bundle", path)
    if not os.path.isfile(full_path):
        raise HTTPException(status_code=404, detail="File not found")
    return FileResponse(full_path)


static_dir = os.path.join(os.path.dirname(__file__), "static")
os.makedirs(static_dir, exist_ok=True)
app.mount("/static", StaticFiles(directory=static_dir), name="static")


@app.get("/")
async def index():
    index_path = os.path.join(static_dir, "index.html")
    if os.path.exists(index_path):
        return FileResponse(index_path)
    return {"message": "Story Config Tool API is running"}
